#! /usr/bin/python

#Generate Word Documents from Word Templates and Excel Source File
import openpyxl
import os
import sys 
import docx

def get_script_path(): #get current directory where script is located
    return os.path.dirname(os.path.realpath(sys.argv[0]))

os.chdir(get_script_path())

#Read source excel file

wb = openpyxl.load_workbook('source.xlsx')
ws = wb['Sheet1']
myList = []
listIndex=0
for row in ws['A{}:E{}'.format(ws.min_row,ws.max_row)]:
    myList.append([])    
    for cell in row:
       myList[listIndex].append(cell.value)
    listIndex += 1
print(myList)
print(myList[0])
print('Source File successfully parsed')
wb.close

textToReplace = myList[0]
for x in myList[1:]:      
    wordDoc = docx.Document('template.docx')
    
    for p in wordDoc.paragraphs: 
        rIndex = 0      
        for tIndex in textToReplace:           
            #print(tIndex)
            #print(rIndex)
            if tIndex in p.text:
                replacementText =   str(x[rIndex]).strip()
                print('found ' + tIndex)                
                inline = p.runs
                # Loop added to work with runs (strings with same style)
                for i in range(len(inline)):
                    if tIndex in inline[i].text:
                        print(replacementText)    
                        text = inline[i].text.replace(tIndex, replacementText.upper())
                        inline[i].text = text    
            rIndex += 1
        
        filename= str(x[0]).strip() + '.docx'
    wordDoc.save(filename)        
   
    print(filename, ' Successfully Created!')
print('Script Finished')